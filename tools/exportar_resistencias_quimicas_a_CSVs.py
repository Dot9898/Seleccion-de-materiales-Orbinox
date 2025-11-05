
from pathlib import Path
import csv
from openpyxl import load_workbook

ROOT_PATH = Path(__file__).resolve().parent.parent
DATA_PATH = ROOT_PATH / 'data'


color_to_quality = {'wh':0, 'oj':1, 'yl':2, 'bl':3, 'gr':4}
rgb_to_color = {'FFFFFFFF':'wh', 'FFF79239':'oj', 'FF8AC53F':'gr', 'FF5CC4EB':'bl', 'FFFFF100':'yl', 'FFE9F6FD':'wh', '00000000':'wh'}

wb = load_workbook(ROOT_PATH / 'data' / 'resistencias_quimicas.xlsx', data_only=True)
ws = wb.active

#arreglar conversión pdf a excel
#arreglar relative paths
with (open(ROOT_PATH / 'data' / 'values.csv', "w", newline="", encoding="utf-8") as f_values, 
      open(ROOT_PATH / 'data' / 'colors.csv', "w", newline="", encoding="utf-8") as f_colors):
    
    writer_values = csv.writer(f_values)
    writer_colors = csv.writer(f_colors)
    
    for row in ws.iter_rows():
        values_row = []
        colors_row = []
        for cell in row:
            # Cell value
            values_row.append(cell.value if cell.value is not None else "")
            
            # Cell color
            fill = cell.fill
            if fill and fill.fgColor.type == "rgb":
                if fill.fgColor.rgb in rgb_to_color:
                    colors_row.append(color_to_quality[rgb_to_color[fill.fgColor.rgb]])
                else:
                    colors_row.append(fill.fgColor.rgb)
            else:
                colors_row.append("None")
        writer_values.writerow(values_row)
        writer_colors.writerow(colors_row)

print("exported")








